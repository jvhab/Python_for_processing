import openpyxl as opyxl                                                    # Импорт библиотеки
wb = opyxl.load_workbook('salaries_6_5_1.xlsx', 'r')                        # Открытие файла xlsx
sh = wb.active                                                              # Выбор листа для работы

# Создание словарей для работы

CityDict = {}
newCityDict = {}
jobDict = {}
newjobDict = {}

# Перебор значений в каждом городе

for city in range(2, sh.max_row + 1):
    City = []
    for salaryInCity in range(1, sh.max_column + 1):
        resultCity = sh.cell(row=city, column=salaryInCity)
        City.append(resultCity.value)                                       # Запись в лист
    CityDict[City[0]] = City[1:]                                            # Запись в словарь, где ключ город

# Перебор значений в словаре CityDict, где i - ключ, k - значение

for i, k in CityDict.items():
    SortSalaryInCity = sorted(k)
    newCityDict[SortSalaryInCity[len(SortSalaryInCity) // 2]] = i           # Ключ в словаре будет значение медианной зарплаты
    mySort = sorted(newCityDict)                                            # Сортировка словоря, последнее значение самое большое (-1 значение)

# Вывод значения и медианной зарплаты

print(newCityDict[mySort[-1]], mySort[-1])

# Перебор зарплат для каждой работы

for job in range(2, sh.max_column + 1):
    Job = []
    for salaryInCity in range(1, sh.max_row + 1):
        resultSalary = sh.cell(row=salaryInCity, column=job)
        Job.append(resultSalary.value)
    jobDict[Job[0]] = Job[1:]

# Перебор в словаре jobDict, где i - название работы, k - зарплата

for i, k in jobDict.items():
    sumSalary = sum(k) // len(k)                                            # Получение средней зарплаты
    newjobDict[sumSalary] = i                                               # Запись в словарь, где ключ средняя зарплата, значение название работы
    jobsort = sorted(newjobDict)                                            # Сортировка словаря, последнее значенние самое большое (-1)

# Вывод самой большой зп и название работы

print(jobsort[-1], newjobDict[jobsort[-1]])

wb.close()