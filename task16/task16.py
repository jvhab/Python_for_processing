import openpyxl as opyxl                                    # Библиотека для работы с xlsx файлами
import math                                                 # Математическая библиотека, взял для округления ответов (math.floor())

wb = opyxl.load_workbook('trekking2_6_6_2.xlsx', 'r')       # Чтение xlsx файла
sh = wb.get_sheet_by_name('Справочник')                     # Выбор рабочего листа

# Переменные
setProduct = {}
setMenu = {}
KKAL = 0
BELOK = 0
FAT = 0
UGLI = 0

# Создание словаря, где ключ - продукт, значение - список из значений ккал, белков, жиров, углеводов
for rows in range(2, sh.max_row + 1):
    Product = []
    for columns in range(1, sh.max_column + 1):
        now = sh.cell(column=columns, row=rows).value
        Product.append(now)
    setProduct[Product[0]] = Product[1:]

# Работа со вторым листом в файле
sh = wb.get_sheet_by_name('Раскладка')

# Создание словаря, где ключ - продукт, значение - граммы продукта
for rows in range(2, sh.max_row + 1):
    Menu = []
    for columns in range(1, sh.max_column + 1):
        now = sh.cell(column=columns, row=rows).value
        Menu.append(now)
    setMenu[Menu[0]] = setMenu.get(Menu[0], 0) + Menu[1]    # Есть повторяющиеся продукты, суммируем их

# Перебор словарей setProduct и setMenu
for j, k in setProduct.items():
    for n, m in setMenu.items():
        if j == n:                                          # Ищем нужный продукт и считаем его ккал, жиры и т.д.
            kal = (k[0] * m) / 100
            bel = (k[1] * m) / 100
            fat = (k[2] * m) / 100
            if k[3] == None:                                # Выдало ошибку, т.к. пропущены некоторые данные (None)
                k[3] = 0                                    # Заменяем их на 0, чтобы не было ошибки
            ugli = (k[3] * m) / 100
            KKAL += kal
            BELOK += bel
            FAT += fat
            UGLI += ugli

# Вывод результатов через пробел
print(math.floor(KKAL), end=' ')
print(math.floor(BELOK), end=' ')
print(math.floor(FAT), end= ' ')
print(math.floor(UGLI))