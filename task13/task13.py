# Импорт библиотек для работы с xlsx-файлами и ОС
import openpyxl as opyxl
import os

# Создание списков для заполнения из xlsx-файлов
myset = []
worker = []

# Перебор файлов из папки
for filename in os.listdir("C:\\Users\Костя\\MyProject\\task13"):
    if filename.endswith(".xlsx"):                                  # Берем файлы с расширением xlsx
        wb = opyxl.load_workbook(filename)                          # Загружаем файл
        sh = wb.active                                              # Берем первый лист из файла
        name = sh.cell(column=2, row=2).value                       # Значение имени
        salary = sh.cell(column=4, row=2).value                     # Значение ЗП
        myset = [name, salary]                                      # Запись в список
        worker.append(myset)                                        # Списки вложил в список
mysort = sorted(worker, key=lambda x: x[0])                         # Сортировка по ФИО, 0 элемент в списке
for i in mysort:                                                    # Перебрал, чтобы убрать []
    print(*i)                                                       # Вывод без []