import openpyxl as opyxl
wb = opyxl.load_workbook('trekking1_6_6_1.xlsx', 'r')
sh = wb.active

Menu = {}


for i in range(2, sh.max_row + 1):                      # Перебор строк из файла
    Product = sh.cell(column=1, row=i).value
    Kkal = sh.cell(column=2, row=i).value
    Menu[Product] = Kkal
sortMenu = sorted(Menu, key=lambda x: (-Menu[x], x))    # Сортировка по значению, затем по названию
for j in sortMenu:                                      # Сортировка для вывода одного значения в строке
    print(j)
