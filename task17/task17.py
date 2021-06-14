# Импорт библиотек
import openpyxl as opyxl
import math

# Открытие файла xlsx и выбор рабочего листа
wb = opyxl.load_workbook('trekking3_6_6_3.xlsx', 'r')
sh = wb['Справочник']

# Создание dict для хранения всего меню и параметров этого меню
# создание list для хранения, что берем в поход по всем дням
setMenu = {}
setDay = []

# Перебор значений в меню и запись всего этого в setMenu
for myStr in range(2, sh.max_row + 1):
    Menu = []
    for myColumn in range(1, sh.max_column + 1):
        now = sh.cell(column=myColumn, row=myStr).value
        if now == None:                                         # Замена значений None на 0
            now = 0
        Menu.append(now)
    setMenu[Menu[0]] = Menu[1:]

# выбор рабочего листа
sh = wb['Раскладка']

# Перебор значений, что берем с собой и запись в SetDay
for myStr in range(2, sh.max_row + 1):
    setOfDay = []
    for myColumn in range(1, sh.max_column + 1):
        now = sh.cell(column=myColumn, row=myStr).value
        setOfDay.append(now)
    setDay.append(setOfDay)

# Вывод итоговых значений
for day in range(1, 10):                                        # Перебор дней от 1 до 9
    KKAL = 0                                                    # Создание нужных переменных для обнуления при новом дне
    BELOK = 0
    FAT = 0
    UGLI = 0
    for i in setDay:                                            # Перебор, что взяли с собой
        if i[0] == day:                                         # Установка какой брать день
            for j, k in setMenu.items():                        # Перебор всего меню по параметрам
                if i[1] == j:                                   # Нахождение совпадения, что взяли с собой и меню
                    kal = (k[0] * i[2]) / 100                   # Подсчет каждого параметра
                    bel = (k[1] * i[2]) / 100
                    fat = (k[2] * i[2]) / 100
                    ugli = (k[3] * i[2]) / 100
                    KKAL += kal                                 # Сумма всей еды за день
                    BELOK += bel
                    FAT += fat
                    UGLI += ugli
    print(math.floor(KKAL), end=' ')                            # Вывод текущий день
    print(math.floor(BELOK), end=' ')
    print(math.floor(FAT), end=' ')
    print(math.floor(UGLI))

